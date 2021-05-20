import requests
import openpyxl
import os.path

# Descarcarea fisierelor excel
if os.path.isfile('../../Desktop/Facultate/an 2/an 2 sem 2/lp2/pro2/transparenta_martie_2021.xlsx'):
    print("Fisierul transparenta_martie_2021.xlsx este deja descarcat.")
else:
    url = "https://data.gov.ro/dataset/b86a78a3-7f88-4b53-a94f-015082592466/resource/d0b60b45-fb08-4980-a34c-8cbb4a43cad3/download/transparenta_martie_2021.xlsx"
    r = requests.get(url, allow_redirects=True)
    open('../../Desktop/Facultate/an 2/an 2 sem 2/lp2/pro2/transparenta_martie_2021.xlsx', 'wb').write(r.content)

if os.path.isfile('../../Desktop/Facultate/an 2/an 2 sem 2/lp2/pro2/transparenta_aprilie_2021.xlsx'):
    print("Fisierul transparenta_aprilie_2021.xlsx este deja descarcat.")
else:
    url2 = "https://data.gov.ro/dataset/b86a78a3-7f88-4b53-a94f-015082592466/resource/d3280256-07cc-4f93-957a-9815085899be/download/transparenta_aprilie_2021.xlsx"
    s = requests.get(url2, allow_redirects=True)
    open('../../Desktop/Facultate/an 2/an 2 sem 2/lp2/pro2/transparenta_aprilie_2021.xlsx', 'wb').write(s.content)

if os.path.isfile('../../Desktop/Facultate/an 2/an 2 sem 2/lp2/pro2/transparenta_mai_2021.xlsx'):
    print("Fisierul transparenta_mai_2021.xlsx este deja descarcat.")
else:
    url3 = "https://data.gov.ro/dataset/b86a78a3-7f88-4b53-a94f-015082592466/resource/12980ac2-f459-4b42-905d-16b93bde62ba/download/transparenta_mai_2021.xlsx"
    t = requests.get(url3, allow_redirects=True)
    open('../../Desktop/Facultate/an 2/an 2 sem 2/lp2/pro2/transparenta_mai_2021.xlsx', 'wb').write(t.content)
print()

# Accesarea fisierelor excel
theFile = openpyxl.load_workbook('transparenta_martie_2021.xlsx')
theFile2 = openpyxl.load_workbook('transparenta_aprilie_2021.xlsx')
theFile3 = openpyxl.load_workbook('transparenta_mai_2021.xlsx')
currentSheet = theFile['incidenta']
currentSheet2 = theFile2['incidenta']
currentSheet3 = theFile3['incidenta']

#Rata de infectare / localitate in ultimele 14 zile + mesaj pentru depasire prag
nrLinii = currentSheet.max_row
nrLinii2 = currentSheet2.max_row
nrLinii3 = currentSheet3.max_row
prag_infectari = 1.0

for i in range(4, nrLinii + 1):
    suma = 0
    for j in 'CDEFGHIJKLMNOP':
        coord = '{}{}'.format(j, i)
        suma += currentSheet[coord].value
        media = suma / 14
    coord_localitate = '{}{}'.format('A', i)
    if media > prag_infectari:
        print("Media infectarii pe 14 zile in " + currentSheet[coord_localitate].value +
          " este " + str(media) + ". Valoarea este peste pragul setat, adica " +
              str(prag_infectari))
    else:
        print("Media infectarii pe 14 zile in " + currentSheet[coord_localitate].value +
              " este " + str(media))
print()
for i in range(4, nrLinii2 + 1):
    suma = 0
    for j in 'CDEFGHIJKLMNOP':
        coord2 = '{}{}'.format(j, i)
        suma += currentSheet[coord2].value
        media = suma / 14
    coord_localitate2 = '{}{}'.format('A', i)
    if media > prag_infectari:
        print("Media infectarii pe 14 zile in " + currentSheet2[coord_localitate2].value +
          " este " + str(media) + ". Valoarea este peste pragul setat, adica " +
              str(prag_infectari))
    else:
        print("Media infectarii pe 14 zile in " + currentSheet2[coord_localitate2].value +
              " este " + str(media))
print()
for i in range(4, nrLinii3 + 1):
    suma = 0
    for j in 'CDEFGHIJKLMNOP':
        coord3 = '{}{}'.format(j, i)
        suma += currentSheet3[coord3].value
        media = suma / 14
    coord_localitate3 = '{}{}'.format('A', i)
    if media > prag_infectari:
        print("Media infectarii pe 14 zile in " + currentSheet3[coord_localitate3].value +
          " este " + str(media) + ". Valoarea este peste pragul setat, adica " +
              str(prag_infectari))
    else:
        print("Media infectarii pe 14 zile in " + currentSheet3[coord_localitate3].value +
              " este " + str(media))
print()

# Afisarea judetelor cu cea mai mare prevalenta in ultimele 10 zile
testare_laborator = theFile['testare_laborator']
rata_pozitiv_laborator = theFile['rata_pozitiv_laborator']
testare_laborator2 = theFile2['testare_laborator']
rata_pozitiv_laborator2 = theFile2['rata_pozitiv_laborator']
testare_laborator3 = theFile3['testare_laborator']
rata_pozitiv_laborator3 = theFile3['rata_pozitiv_laborator']

nrMaxLinii =  testare_laborator.max_row
nrMaxLinii2 =  testare_laborator2.max_row
nrMaxLinii3 =  testare_laborator3.max_row

maxim_infectari = 0
maxim_infectari_judet = ''
maxim_infectari2 = 0
maxim_infectari_judet2 = ''
maxim_infectari3 = 0
maxim_infectari_judet3 = ''

dictionar_judet_infectari = {}
dictionar_judet_infectari2 = {}
dictionar_judet_infectari3 = {}

for i in range(4, nrMaxLinii + 1):
    suma = 0
    for j in 'BCDEFGHIJK':
        coord2 = '{}{}'.format(j, i)
        nume_judet = '{}{}'.format('A', i)
        if type(testare_laborator['{}{}'.format(j, i)].value) != int:
            nr_teste = 0
        else:
            nr_teste = testare_laborator['{}{}'.format(j, i)].value
        if type(rata_pozitiv_laborator['{}{}'.format(j, i)].value) != float:
                rata_pozitiv = 0
        else:
                rata_pozitiv = rata_pozitiv_laborator['{}{}'.format(j, i)].value
        pozitivi_pe_zi = rata_pozitiv * nr_teste
        suma += round(pozitivi_pe_zi)
    print("Judetul " + testare_laborator[nume_judet].value + " are in luna martie " +
          str(suma) + " infectati in ultimele 10 zile.")
    dictionar_judet_infectari[suma]= testare_laborator[nume_judet].value
    if suma > maxim_infectari:
        maxim_infectari = suma
        maxim_infectari_judet = testare_laborator[nume_judet].value

for i in range(4, nrMaxLinii2 + 1):
    suma2 = 0

    for j in 'BCDEFGHIJK':
        coord2 = '{}{}'.format(j, i)
        nume_judet2 = '{}{}'.format('A', i)
        if type(testare_laborator2['{}{}'.format(j, i)].value) != int:
            nr_teste2 = 0
        else:
            nr_teste2 = testare_laborator2['{}{}'.format(j, i)].value
        if type(rata_pozitiv_laborator2['{}{}'.format(j, i)].value) != float:
            rata_pozitiv2 = 0
        else:
            rata_pozitiv2 = rata_pozitiv_laborator2['{}{}'.format(j, i)].value
        pozitivi_pe_zi2 = rata_pozitiv2 * nr_teste2
        suma2 += round(pozitivi_pe_zi2)
    print("Judetul " + testare_laborator2[nume_judet2].value + " are in luna aprilie " +
              str(suma2) + " infectati in ultimele 10 zile.")
    dictionar_judet_infectari2[suma2] = testare_laborator2[nume_judet2].value
    if suma2 > maxim_infectari2:
        maxim_infectari2 = suma2
        maxim_infectari_judet2 = testare_laborator2[nume_judet2].value

for i in range(4, nrMaxLinii3 + 1):
    suma3 = 0

    for j in 'BCDEFGHIJK':
        coord3 = '{}{}'.format(j, i)
        nume_judet3 = '{}{}'.format('A', i)
        if type(testare_laborator3['{}{}'.format(j, i)].value) != int:
            nr_teste3 = 0
        else:
            nr_teste3 = testare_laborator3['{}{}'.format(j, i)].value
        if type(rata_pozitiv_laborator3['{}{}'.format(j, i)].value) != float:
            rata_pozitiv3 = 0
        else:
            rata_pozitiv3 = rata_pozitiv_laborator3['{}{}'.format(j, i)].value
        pozitivi_pe_zi3 = rata_pozitiv3 * nr_teste3
        suma3 += round(pozitivi_pe_zi3)
    print("Judetul " + testare_laborator3[nume_judet3].value + " are in luna mai " +
      str(suma3) + " infectati in ultimele 10 zile.")
    dictionar_judet_infectari3[suma3] = testare_laborator3[nume_judet3].value
    if suma3 > maxim_infectari3:
        maxim_infectari3 = suma3
        maxim_infectari_judet3 = testare_laborator3[nume_judet3].value

print()
print("Numarul maxim de infectari pe 10 zile(martie): " + str(maxim_infectari))
print("Judetul/Judetele cu numarul maxim de infectari pe 10 zile (martie) este/sunt: " +
      str(dictionar_judet_infectari.get(maxim_infectari)))
print()
print("Numarul maxim de infectari pe 10 zile (aprilie): " + str(maxim_infectari2))
print("Judetul/Judetele cu numarul maxim de infectari pe 10  zile (aprilie) este/sunt: " +
      str(dictionar_judet_infectari2.get(maxim_infectari2)))
print()
print("Numarul maxim de infectari pe 10 zile(mai): " + str(maxim_infectari3))
print("Judetul/Judetele cu numarul maxim de infectari pe 10 zile (mai) este/sunt: " +
      str(dictionar_judet_infectari3.get(maxim_infectari3)))









